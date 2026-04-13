"""
PassMark BurnInTest 磁碟讀寫速度 + 溫度監控程式 v1.6
────────────────────────────────────────────────────
磁碟速度擷取：
  - 優先使用 pywinauto 直接讀取 BurnInTest 視窗文字
  - 備用：OCR 截圖（原圖放大 4x + UnsharpMask）
  - 補值邏輯：切換 Writing/Verifying 時，另一欄保留上一筆數值

外殼溫度（序列埠）：
  - 通訊協定：baudrate=9600，指令 0xAA 0x55 0x01 0x03 0x03
  - 回應封包第 4~5 byte (little-endian) ÷ 10 = 外殼溫度 (°C)
  - 與磁碟速度同步採樣，記錄到同一列

SMART 溫度（USB SSD 直讀）：
  - 透過 Win32 DeviceIoControl 讀取 USB SSD 的 SMART 溫度
  - 優先嘗試 StorageDeviceTemperatureProperty（Win10+ 原生）
  - 備用：SCSI ATA PASS-THROUGH(12)（SM2322 / UAS / UASP 裝置）
  - 備用：IOCTL_ATA_PASS_THROUGH（舊式橋接晶片）
  - 需要以系統管理員身分執行，啟動時自動 UAC 提升

依賴套件：
    pip install pillow pytesseract mss openpyxl pywinauto pywin32 pyserial
    另需安裝 Tesseract OCR：https://github.com/UB-Mannheim/tesseract/wiki
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
import re
import datetime
import os
import sys
import ctypes
import ctypes.wintypes as wt
import struct
from dataclasses import dataclass
from typing import Optional

# ══════════════════════════════════════════════════════════════════
#  UAC 管理員提升（SMART 溫度讀取需要）
# ══════════════════════════════════════════════════════════════════

def _is_admin() -> bool:
    try:
        return bool(ctypes.windll.shell32.IsUserAnAdmin())
    except Exception:
        return False


def _relaunch_as_admin() -> None:
    """靜默以管理員身分重新執行本程式。"""
    ctypes.windll.shell32.ShellExecuteW(
        None, "runas", sys.executable,
        " ".join(f'"{a}"' for a in sys.argv),
        None, 1,
    )


# ── 延遲匯入 ────────────────────────────────────────────────────
MISSING = []
try:
    import mss as mss_module
except ImportError:
    MISSING.append("mss")
    mss_module = None

try:
    from PIL import Image, ImageEnhance, ImageOps, ImageFilter
except ImportError:
    MISSING.append("Pillow")

try:
    import pytesseract
except ImportError:
    MISSING.append("pytesseract")

try:
    import openpyxl
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    MISSING.append("openpyxl")

# pyserial 為可選（外殼溫度功能需要）
try:
    import serial
    from serial.tools import list_ports
    HAS_SERIAL = True
except ImportError:
    HAS_SERIAL = False

# pywinauto 為可選（視窗直讀需要）
try:
    import pywinauto
    from pywinauto import Application as WinApp
    HAS_PYWINAUTO = True
except ImportError:
    HAS_PYWINAUTO = False

# ── Tesseract 路徑 ───────────────────────────────────────────────
def _get_tesseract_search_paths():
    paths = []
    # 1. 若為 PyInstaller 打包的 .exe，優先找內嵌的 Tesseract-OCR 資料夾
    if getattr(sys, 'frozen', False):
        embedded = os.path.join(sys._MEIPASS, "Tesseract-OCR", "tesseract.exe")
        paths.append(embedded)
    # 2. 系統安裝路徑
    paths += [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        r"C:\Users\{}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe".format(
            os.environ.get("USERNAME", "")
        ),
    ]
    return paths

TESSERACT_DEFAULT_PATHS = _get_tesseract_search_paths()

# BurnInTest 視窗標題關鍵字
BURNIN_WINDOW_KEYWORDS = ["BurnInTest", "Burn-In Test", "PassMark"]

# 外殼溫度感測器通訊協定常數
TEMP_BAUDRATE = 9600
TEMP_REQUEST  = bytes([0xAA, 0x55, 0x01, 0x03, 0x03])
TEMP_RESP_LEN = 13


# ══════════════════════════════════════════════════════════════════
#  外殼溫度感測器讀取模組（序列埠）
# ══════════════════════════════════════════════════════════════════
class TemperatureReader:
    """
    封裝外殼溫度感測器的序列埠通訊。
    協定：送出 TEMP_REQUEST (5 bytes)，讀回 13 bytes，
          取第 4~5 byte (little-endian) ÷ 10 = 外殼溫度 (°C)
    """

    def __init__(self):
        self._port   = None   # serial.Serial 物件
        self._lock   = threading.Lock()
        self.enabled = False  # 由 GUI 控制是否啟用

    @staticmethod
    def list_ports():
        """回傳可用 COM port 列表"""
        if not HAS_SERIAL:
            return []
        try:
            return [p.device for p in list_ports.comports()]
        except Exception:
            return []

    def open(self, port_name):
        """開啟序列埠，失敗拋出 Exception"""
        if not HAS_SERIAL:
            raise RuntimeError("pyserial 未安裝，請執行: pip install pyserial")
        # 先關閉舊的埠（不持鎖，避免死鎖）
        self._close_port_nolock()
        with self._lock:
            self._port = serial.Serial(
                port=port_name,
                baudrate=TEMP_BAUDRATE,
                timeout=2
            )
            self.enabled = True

    def _close_port_nolock(self):
        """不取鎖的內部關閉（供 open() 安全呼叫）"""
        self.enabled = False
        try:
            if self._port and self._port.is_open:
                self._port.close()
        except Exception:
            pass
        self._port = None

    def close(self):
        """關閉序列埠（外部呼叫用，有鎖保護）"""
        with self._lock:
            self._close_port_nolock()

    def read_temperature(self):
        """
        送出指令並讀回外殼溫度。
        回傳 float (°C) 或 None（失敗/未啟用）
        """
        if not self.enabled:
            return None
        with self._lock:
            if not self._port or not self._port.is_open:
                return None
            try:
                self._port.write(TEMP_REQUEST)
                response = self._port.read(TEMP_RESP_LEN)
                if len(response) < 6:
                    return None
                t1 = int.from_bytes(response[4:6], "little") / 10.0
                return t1
            except Exception:
                return None


# ══════════════════════════════════════════════════════════════════
#  SMART 溫度讀取模組（USB SSD / Win32 DeviceIoControl）
#  移植自 smart_reader.py，整合為單一檔案，不需外部依賴
# ══════════════════════════════════════════════════════════════════

# Win32 常數
_GENERIC_READ          = 0x80000000
_GENERIC_WRITE         = 0x40000000
_FILE_SHARE_READ       = 0x00000001
_FILE_SHARE_WRITE      = 0x00000002
_OPEN_EXISTING         = 3
_FILE_ATTRIBUTE_NORMAL = 0x80
_INVALID_HANDLE_VALUE  = ctypes.c_void_p(-1).value

# IOCTL codes
_IOCTL_STORAGE_QUERY_PROPERTY = 0x002D1400
_IOCTL_SCSI_PASS_THROUGH      = 0x0004D004
_IOCTL_ATA_PASS_THROUGH       = 0x0004D02C

# STORAGE_PROPERTY_ID
_StorageDeviceProperty            = 0
_StorageDeviceTemperatureProperty = 53
_PropertyStandardQuery            = 0

# SCSI
_SCSI_IOCTL_DATA_IN = 1
_SENSE_BUF_SIZE     = 32
_DATA_BUF_SIZE      = 512

# ATA SMART
_SMART_CMD          = 0xB0
_SMART_READ_ATTRIBS = 0xD0
_SMART_CYL_LOW      = 0x4F
_SMART_CYL_HIGH     = 0xC2
_SMART_DEVICE       = 0xA0
_ATA_FLAGS_DATA_IN  = 0x0002

_BUS_TYPE_NAMES = {1:"SCSI",2:"ATAPI",3:"ATA",7:"USB",11:"SATA",17:"NVMe"}
_BusTypeUsb = 7

# ctypes 結構
class _STORAGE_PROPERTY_QUERY(ctypes.Structure):
    _fields_ = [("PropertyId",ctypes.c_int),("QueryType",ctypes.c_int),
                ("AdditionalParameters",ctypes.c_ubyte*1)]

class _STORAGE_TEMPERATURE_INFO(ctypes.Structure):
    _fields_ = [("Index",ctypes.c_ushort),("Temperature",ctypes.c_short),
                ("OverThreshold",ctypes.c_short),("UnderThreshold",ctypes.c_short),
                ("OverThresholdChangable",ctypes.c_ubyte),
                ("UnderThresholdChangable",ctypes.c_ubyte),
                ("EventGenerated",ctypes.c_ubyte),("Reserved0",ctypes.c_ubyte),
                ("Reserved1",ctypes.c_ulong)]

class _STORAGE_TEMPERATURE_DATA_DESCRIPTOR(ctypes.Structure):
    _fields_ = [("Version",ctypes.c_ulong),("Size",ctypes.c_ulong),
                ("CriticalTemperature",ctypes.c_short),
                ("WarningTemperature",ctypes.c_short),
                ("InfoCount",ctypes.c_ushort),("Reserved0",ctypes.c_ubyte*2)]

class _STORAGE_DEVICE_DESCRIPTOR_HEADER(ctypes.Structure):
    _fields_ = [("Version",ctypes.c_ulong),("Size",ctypes.c_ulong),
                ("DeviceType",ctypes.c_ubyte),("DeviceTypeModifier",ctypes.c_ubyte),
                ("RemovableMedia",ctypes.c_ubyte),("CommandQueueing",ctypes.c_ubyte),
                ("VendorIdOffset",ctypes.c_ulong),("ProductIdOffset",ctypes.c_ulong),
                ("ProductRevisionOffset",ctypes.c_ulong),
                ("SerialNumberOffset",ctypes.c_ulong),
                ("BusType",ctypes.c_ubyte),("RawPropertiesLength",ctypes.c_ulong)]

class _SCSI_PASS_THROUGH(ctypes.Structure):
    _fields_ = [("Length",ctypes.c_ushort),("ScsiStatus",ctypes.c_ubyte),
                ("PathId",ctypes.c_ubyte),("TargetId",ctypes.c_ubyte),
                ("Lun",ctypes.c_ubyte),("CdbLength",ctypes.c_ubyte),
                ("SenseInfoLength",ctypes.c_ubyte),("DataIn",ctypes.c_ubyte),
                ("_pad",ctypes.c_ubyte*3),("DataTransferLength",ctypes.c_ulong),
                ("TimeOutValue",ctypes.c_ulong),("DataBufferOffset",ctypes.c_size_t),
                ("SenseInfoOffset",ctypes.c_ulong),("Cdb",ctypes.c_ubyte*16)]

class _SCSI_PASS_THROUGH_WITH_BUFFERS(ctypes.Structure):
    _fields_ = [("Spt",_SCSI_PASS_THROUGH),
                ("SenseBuf",ctypes.c_ubyte*_SENSE_BUF_SIZE),
                ("DataBuf",ctypes.c_ubyte*_DATA_BUF_SIZE)]

class _ATA_PASS_THROUGH_EX(ctypes.Structure):
    _pack_ = 1
    _fields_ = [("Length",ctypes.c_ushort),("AtaFlags",ctypes.c_ushort),
                ("PathId",ctypes.c_ubyte),("TargetId",ctypes.c_ubyte),
                ("Lun",ctypes.c_ubyte),("ReservedAsUchar",ctypes.c_ubyte),
                ("DataTransferLength",ctypes.c_ulong),("TimeOutValue",ctypes.c_ulong),
                ("ReservedAsUlong",ctypes.c_ulong),("DataBufferOffset",ctypes.c_size_t),
                ("PreviousTaskFile",ctypes.c_ubyte*8),("CurrentTaskFile",ctypes.c_ubyte*8)]

# Win32 API 綁定
try:
    _k32 = ctypes.windll.kernel32
    _k32.CreateFileW.restype  = ctypes.c_void_p
    _k32.CreateFileW.argtypes = [wt.LPCWSTR,wt.DWORD,wt.DWORD,ctypes.c_void_p,
                                  wt.DWORD,wt.DWORD,ctypes.c_void_p]
    _k32.DeviceIoControl.restype  = wt.BOOL
    _k32.DeviceIoControl.argtypes = [ctypes.c_void_p,wt.DWORD,
                                      ctypes.c_void_p,wt.DWORD,
                                      ctypes.c_void_p,wt.DWORD,
                                      ctypes.POINTER(wt.DWORD),ctypes.c_void_p]
    _k32.CloseHandle.restype  = wt.BOOL
    _k32.CloseHandle.argtypes = [ctypes.c_void_p]
    HAS_SMART = True
except Exception:
    HAS_SMART = False


@dataclass
class DriveInfo:
    device_path: str
    model:       str  = "Unknown"
    is_usb:      bool = False
    bus_type:    str  = ""
    temperature: Optional[int] = None
    method:      str  = ""


def _open_drive(path: str):
    if not HAS_SMART:
        return None
    h = _k32.CreateFileW(path, _GENERIC_READ | _GENERIC_WRITE,
                          _FILE_SHARE_READ | _FILE_SHARE_WRITE,
                          None, _OPEN_EXISTING, _FILE_ATTRIBUTE_NORMAL, None)
    return None if (h == _INVALID_HANDLE_VALUE or h is None) else h


def _ioctl(handle, code, in_ptr, in_sz, out_ptr, out_sz):
    returned = wt.DWORD(0)
    ok = _k32.DeviceIoControl(handle, code,
                               ctypes.cast(in_ptr,  ctypes.c_void_p), in_sz,
                               ctypes.cast(out_ptr, ctypes.c_void_p), out_sz,
                               ctypes.byref(returned), None)
    return returned.value if ok else None


def _read_ascii(buf: bytes, offset: int) -> str:
    if offset == 0 or offset >= len(buf):
        return ""
    end = buf.find(b"\x00", offset)
    end = end if end != -1 else len(buf)
    return buf[offset:end].decode("ascii", errors="replace").strip()


def _fill_descriptor(handle, info: DriveInfo) -> None:
    buf_sz = 1024
    buf    = (ctypes.c_ubyte * buf_sz)()
    q      = _STORAGE_PROPERTY_QUERY(PropertyId=_StorageDeviceProperty,
                                      QueryType=_PropertyStandardQuery)
    ret = _ioctl(handle, _IOCTL_STORAGE_QUERY_PROPERTY,
                 ctypes.byref(q), ctypes.sizeof(q), buf, buf_sz)
    if ret is None:
        return
    raw  = bytes(buf)
    desc = _STORAGE_DEVICE_DESCRIPTOR_HEADER.from_buffer_copy(raw)
    info.bus_type = _BUS_TYPE_NAMES.get(desc.BusType, f"Type({desc.BusType})")
    info.is_usb   = (desc.BusType == _BusTypeUsb)
    model = _read_ascii(raw, desc.ProductIdOffset)
    if not model:
        model = _read_ascii(raw, desc.VendorIdOffset)
    if model:
        info.model = model


def _query_temp_property(handle) -> Optional[int]:
    desc_sz = ctypes.sizeof(_STORAGE_TEMPERATURE_DATA_DESCRIPTOR)
    info_sz = ctypes.sizeof(_STORAGE_TEMPERATURE_INFO)
    buf_sz  = desc_sz + 16 * info_sz
    q   = _STORAGE_PROPERTY_QUERY(PropertyId=_StorageDeviceTemperatureProperty,
                                   QueryType=_PropertyStandardQuery)
    out = (ctypes.c_ubyte * buf_sz)()
    ret = _ioctl(handle, _IOCTL_STORAGE_QUERY_PROPERTY,
                 ctypes.byref(q), ctypes.sizeof(q), out, buf_sz)
    if ret is None or ret < desc_sz:
        return None
    desc = _STORAGE_TEMPERATURE_DATA_DESCRIPTOR.from_buffer_copy(bytes(out))
    if desc.InfoCount == 0:
        return None
    ti = _STORAGE_TEMPERATURE_INFO.from_buffer_copy(
        bytes(out)[desc_sz: desc_sz + info_sz])
    t = int(ti.Temperature)
    return t if -40 <= t <= 120 else None


def _query_smart_scsi_sat(handle) -> Optional[int]:
    sptwb = _SCSI_PASS_THROUGH_WITH_BUFFERS()
    sz    = ctypes.sizeof(_SCSI_PASS_THROUGH_WITH_BUFFERS)
    spt   = sptwb.Spt
    spt.Length             = ctypes.sizeof(_SCSI_PASS_THROUGH)
    spt.CdbLength          = 12
    spt.SenseInfoLength    = _SENSE_BUF_SIZE
    spt.DataIn             = _SCSI_IOCTL_DATA_IN
    spt.DataTransferLength = _DATA_BUF_SIZE
    spt.TimeOutValue       = 2
    spt.DataBufferOffset   = ctypes.sizeof(_SCSI_PASS_THROUGH) + _SENSE_BUF_SIZE
    spt.SenseInfoOffset    = ctypes.sizeof(_SCSI_PASS_THROUGH)
    cdb = spt.Cdb
    cdb[0] = 0xA1
    cdb[1] = (4 << 1) | 0
    cdb[2] = (1 << 3) | (1 << 2) | 2
    cdb[3] = _SMART_READ_ATTRIBS
    cdb[4] = 1;  cdb[5] = 1
    cdb[6] = _SMART_CYL_LOW;  cdb[7] = _SMART_CYL_HIGH
    cdb[8] = _SMART_DEVICE;   cdb[9] = _SMART_CMD
    ret = _ioctl(handle, _IOCTL_SCSI_PASS_THROUGH,
                 ctypes.byref(sptwb), sz, ctypes.byref(sptwb), sz)
    if ret is None:
        return None
    return _parse_smart_temp(bytes(sptwb.DataBuf))


def _query_smart_ata_direct(handle) -> Optional[int]:
    apt_sz = ctypes.sizeof(_ATA_PASS_THROUGH_EX)
    total  = apt_sz + _DATA_BUF_SIZE
    buf    = (ctypes.c_ubyte * total)()
    apt    = _ATA_PASS_THROUGH_EX()
    apt.Length             = apt_sz
    apt.AtaFlags           = _ATA_FLAGS_DATA_IN
    apt.DataTransferLength = _DATA_BUF_SIZE
    apt.TimeOutValue       = 3
    apt.DataBufferOffset   = apt_sz
    tf    = (ctypes.c_ubyte * 8)()
    tf[0] = _SMART_READ_ATTRIBS
    tf[3] = _SMART_CYL_LOW;  tf[4] = _SMART_CYL_HIGH
    tf[5] = _SMART_DEVICE;   tf[6] = _SMART_CMD
    apt.CurrentTaskFile = tf
    ctypes.memmove(buf, ctypes.byref(apt), apt_sz)
    ret = _ioctl(handle, _IOCTL_ATA_PASS_THROUGH, buf, total, buf, total)
    if ret is None:
        return None
    return _parse_smart_temp(bytes(buf)[apt_sz:])


def _parse_smart_temp(data: bytes) -> Optional[int]:
    temp_c2: Optional[int] = None
    temp_be: Optional[int] = None
    for i in range(30):
        off = 2 + i * 12
        if off + 12 > len(data):
            break
        attr_id = data[off]
        if attr_id == 0:
            continue
        raw_byte0 = data[off + 5]
        if attr_id == 0xC2:
            temp_c2 = raw_byte0
        elif attr_id == 0xBE:
            temp_be = raw_byte0
    result = temp_c2 if temp_c2 is not None else temp_be
    if result is not None and 0 <= result <= 120:
        return result
    return None


def enumerate_usb_drives() -> list:
    """掃描 PhysicalDrive0~15，只回傳 USB 外接裝置。"""
    if not HAS_SMART:
        return []
    drives = []
    for i in range(16):
        path = f"\\\\.\\PhysicalDrive{i}"
        h    = _open_drive(path)
        if h is None:
            continue
        info = DriveInfo(device_path=path)
        _fill_descriptor(h, info)
        _k32.CloseHandle(h)
        if info.is_usb:
            drives.append(info)
    return drives


def refresh_smart_temperature(info: DriveInfo) -> None:
    """依序嘗試三種方式讀取 SMART 溫度，結果寫入 info.temperature / info.method。"""
    h = _open_drive(info.device_path)
    if h is None:
        info.temperature = None
        info.method      = "無法開啟裝置（需管理員權限）"
        return
    try:
        t = _query_temp_property(h)
        if t is not None:
            info.temperature = t
            info.method      = "StorageDeviceTemperatureProperty"
            return
        t = _query_smart_scsi_sat(h)
        if t is not None:
            info.temperature = t
            info.method      = "SCSI ATA PASS-THROUGH(12) [SAT]"
            return
        t = _query_smart_ata_direct(h)
        if t is not None:
            info.temperature = t
            info.method      = "ATA PASS-THROUGH [Direct]"
            return
        info.temperature = None
        info.method      = "橋接晶片不支援溫度讀取"
    finally:
        _k32.CloseHandle(h)


# ══════════════════════════════════════════════════════════════════
#  SMART 溫度管理器（背景定期掃描，供 MonitorThread 呼叫）
# ══════════════════════════════════════════════════════════════════
class SmartTempManager:
    """
    在背景執行緒定期掃描 USB SSD，供 MonitorThread 同步讀取。
    - enabled：由 GUI 控制是否啟用
    - selected_drive：目前選定的 DriveInfo（None = 未選定）
    - last_temp：最後讀到的溫度（°C 整數或 None）
    """

    def __init__(self):
        self.enabled        = False
        self.selected_drive: Optional[DriveInfo] = None
        self.last_temp:      Optional[int]        = None
        self._lock          = threading.Lock()

    def scan_drives(self) -> list:
        """掃描並回傳 USB 裝置清單（在背景執行緒執行）"""
        return enumerate_usb_drives()

    def select_drive(self, drive: Optional[DriveInfo]):
        with self._lock:
            self.selected_drive = drive
            self.last_temp = None

    def read_temperature(self) -> Optional[int]:
        """MonitorThread 每個採樣週期呼叫一次，非阻塞快速讀取。"""
        if not self.enabled:
            return None
        with self._lock:
            d = self.selected_drive
        if d is None:
            return None
        try:
            refresh_smart_temperature(d)
            with self._lock:
                self.last_temp = d.temperature
            return d.temperature
        except Exception:
            return None


# ══════════════════════════════════════════════════════════════════
#  pywinauto 直接讀取 BurnInTest 視窗（優先）
# ══════════════════════════════════════════════════════════════════
class WindowTextReader:
    def __init__(self):
        pass

    def _find_window(self):
        if not HAS_PYWINAUTO:
            return None
        try:
            import win32gui  # type: ignore[import-not-found]
            result = {"hwnd": None}

            def enum_cb(hwnd, _):
                title = win32gui.GetWindowText(hwnd)
                if any(k.lower() in title.lower() for k in BURNIN_WINDOW_KEYWORDS):
                    if win32gui.IsWindowVisible(hwnd):
                        result["hwnd"] = hwnd
                        return False
                return True

            win32gui.EnumWindows(enum_cb, None)
            if result["hwnd"]:
                app = WinApp(backend="win32").connect(handle=result["hwnd"])
                return app.window(handle=result["hwnd"])
        except Exception:
            pass
        return None

    def _collect_text(self, window):
        texts = []
        try:
            for ctrl in window.descendants():
                try:
                    t = ctrl.window_text().strip()
                    if t:
                        texts.append(t)
                except Exception:
                    pass
        except Exception:
            pass
        return "\n".join(texts)

    def read(self) -> Optional[str]:
        w = self._find_window()
        if w is None:
            return None
        return self._collect_text(w)


# ══════════════════════════════════════════════════════════════════
#  OCR 截圖模組（備用）
# ══════════════════════════════════════════════════════════════════
class ScreenCapture:
    def __init__(self, monitor_index=1):
        self.monitor_index = monitor_index

    def grab_region(self, x, y, width, height):
        with mss_module.mss() as sct:
            raw = sct.grab({"top": y, "left": x,
                            "width": width, "height": height})
            return Image.frombytes("RGB", raw.size, raw.bgra, "raw", "BGRX")

    def grab_full_screen(self):
        with mss_module.mss() as sct:
            monitors = sct.monitors
            idx = min(self.monitor_index, len(monitors) - 1)
            raw = sct.grab(monitors[idx])
            return Image.frombytes("RGB", raw.size, raw.bgra, "raw", "BGRX")

    @staticmethod
    def preprocess_best(img):
        w, h = img.size
        img = img.resize((w * 4, h * 4), Image.LANCZOS)
        img = img.filter(ImageFilter.UnsharpMask(radius=2, percent=150, threshold=3))
        return img

    def capture_and_ocr(self, region=None) -> str:
        if region:
            img = self.grab_region(region["x"], region["y"],
                                   region["w"], region["h"])
        else:
            img = self.grab_full_screen()
        processed = self.preprocess_best(img)
        return pytesseract.image_to_string(processed, config="--oem 3 --psm 6")


# ══════════════════════════════════════════════════════════════════
#  速度解析
# ══════════════════════════════════════════════════════════════════
class SpeedParser:
    _RE_SPEED = re.compile(
        r"[Cc]urrent\s*[Ss]peed\s*[:\-\s]+([0-9]+\.?[0-9]*)\s*MB[/\s]?[Ss]ec",
        re.IGNORECASE
    )
    _RE_STATE = re.compile(r"\b(Writing|Verifying)\b", re.IGNORECASE)

    @classmethod
    def parse(cls, text: str) -> dict:
        result = {"status": "unknown", "read_speed": None, "write_speed": None}
        if not text:
            return result
        state_m = cls._RE_STATE.search(text)
        speed_m = cls._RE_SPEED.search(text)
        if not state_m or not speed_m:
            return result
        state = state_m.group(1).lower()
        speed = float(speed_m.group(1))
        if state == "writing":
            result["status"]      = "writing"
            result["write_speed"] = speed
        elif state == "verifying":
            result["status"]     = "verifying"
            result["read_speed"] = speed
        return result


# ══════════════════════════════════════════════════════════════════
#  滑鼠拖拉區域選取
# ══════════════════════════════════════════════════════════════════
class RegionSelector:
    def __init__(self, parent):
        self.parent = parent
        self.result = None

    def select(self):
        overlay = tk.Toplevel(self.parent)
        try:
            import win32api  # type: ignore[import-not-found]
            left = win32api.GetSystemMetrics(76)
            top  = win32api.GetSystemMetrics(77)
            w    = win32api.GetSystemMetrics(78)
            h    = win32api.GetSystemMetrics(79)
        except Exception:
            left = 0
            top  = 0
            w    = overlay.winfo_screenwidth()
            h    = overlay.winfo_screenheight()

        overlay.geometry(f"{w}x{h}+{left}+{top}")
        overlay.overrideredirect(True)
        overlay.attributes("-topmost", True)
        overlay.attributes("-alpha", 0.30)
        overlay.configure(bg="black")
        overlay.config(cursor="crosshair")

        canvas = tk.Canvas(overlay, bg="black", highlightthickness=0,
                           cursor="crosshair")
        canvas.pack(fill="both", expand=True)

        self._start_x = self._start_y = 0
        rect_id = None
        lbl_id  = None

        def on_press(e):
            nonlocal rect_id, lbl_id
            self._start_x, self._start_y = e.x, e.y
            if rect_id:
                canvas.delete(rect_id)
            if lbl_id:
                canvas.delete(lbl_id)
            rect_id = canvas.create_rectangle(e.x, e.y, e.x, e.y,
                                               outline="lime", width=2)

        def on_drag(e):
            nonlocal lbl_id
            canvas.coords(rect_id, self._start_x, self._start_y, e.x, e.y)
            rw = abs(e.x - self._start_x)
            rh = abs(e.y - self._start_y)
            if lbl_id:
                canvas.delete(lbl_id)
            lbl_id = canvas.create_text(e.x + 10, e.y + 10,
                                        text=f"{rw}×{rh}",
                                        fill="lime", font=("Consolas", 11, "bold"),
                                        anchor="nw")

        def on_release(e):
            x1, y1 = min(self._start_x, e.x), min(self._start_y, e.y)
            x2, y2 = max(self._start_x, e.x), max(self._start_y, e.y)
            rw, rh = x2 - x1, y2 - y1
            self.result = {"x": x1+left, "y": y1+top,
                           "w": rw, "h": rh} if rw > 10 and rh > 10 else None
            overlay.destroy()

        def on_esc(e):
            self.result = None
            overlay.destroy()

        canvas.bind("<ButtonPress-1>",   on_press)
        canvas.bind("<B1-Motion>",       on_drag)
        canvas.bind("<ButtonRelease-1>", on_release)
        overlay.bind("<Escape>",         on_esc)
        self.parent.wait_window(overlay)
        return self.result


# ══════════════════════════════════════════════════════════════════
#  監控核心（磁碟速度 + 外殼溫度 + SMART 溫度同步）
# ══════════════════════════════════════════════════════════════════
class Monitor:
    def __init__(self, region=None):
        self.region         = region
        self._win_reader    = WindowTextReader()
        self._screen_reader = ScreenCapture()
        self._use_window    = HAS_PYWINAUTO

    def capture_disk(self):
        """回傳 (speed_dict, using_window)"""
        text = None
        if self._use_window:
            try:
                text = self._win_reader.read()
            except Exception:
                text = None
        if not text:
            self._use_window = False
            try:
                text = self._screen_reader.capture_and_ocr(self.region)
            except Exception:
                text = ""
        return SpeedParser.parse(text), self._use_window

    @property
    def mode(self):
        return "視窗直讀" if self._use_window else "OCR 截圖"


# ══════════════════════════════════════════════════════════════════
#  監控執行緒（磁碟速度 + 外殼溫度 + SMART 溫度同步採樣）
# ══════════════════════════════════════════════════════════════════
class MonitorThread(threading.Thread):
    def __init__(self, interval_sec, total_sec, region,
                 temp_reader, smart_manager,
                 on_record, on_finish, on_error, on_mode_update):
        super().__init__(daemon=True)
        self.interval_sec   = interval_sec
        self.total_sec      = total_sec
        self.temp_reader    = temp_reader    # TemperatureReader（外殼溫度）
        self.smart_manager  = smart_manager  # SmartTempManager（SMART 溫度）
        self.on_record      = on_record
        self.on_finish      = on_finish
        self.on_error       = on_error
        self.on_mode_update = on_mode_update
        self._stop_event    = threading.Event()
        self._monitor       = Monitor(region)

    def stop(self):
        self._stop_event.set()

    def run(self):
        start     = time.time()
        last_mode = None
        while not self._stop_event.is_set():
            if self.total_sec > 0 and (time.time() - start) >= self.total_sec:
                break
            try:
                # 同步讀取磁碟速度 + 外殼溫度 + SMART 溫度
                data, using_window = self._monitor.capture_disk()
                data["temperature"]      = self.temp_reader.read_temperature()
                data["smart_temp"]       = self.smart_manager.read_temperature()
                ts = datetime.datetime.now()
                self.on_record(ts, data)

                mode = "視窗直讀 ✓" if using_window else "OCR 截圖"
                if mode != last_mode:
                    last_mode = mode
                    self.on_mode_update(mode)
            except Exception as e:
                self.on_error(str(e))
            self._stop_event.wait(self.interval_sec)
        self.on_finish()


# ══════════════════════════════════════════════════════════════════
#  Excel 匯出（外殼溫度 + SMART 溫度兩欄及對應圖表）
# ══════════════════════════════════════════════════════════════════
class ExcelExporter:
    HDR_COLOR        = "2F5496"
    READ_COLOR       = "4472C4"   # 讀取：藍
    WRITE_COLOR      = "ED7D31"   # 寫入：橘
    SHELL_TEMP_COLOR = "C00000"   # 外殼溫度：紅
    SMART_TEMP_COLOR = "7030A0"   # SMART 溫度：紫
    TIME_COL         = 8          # H 欄：累積時間（秒）

    def export(self, records, output_path, interval_sec):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "速度與溫度記錄"
        self._header(ws, interval_sec)
        self._data(ws, records, interval_sec)
        self._chart_speed(wb, ws, records)
        self._chart_temp(wb, ws, records)
        self._widths(ws)
        wb.save(output_path)

    def _header(self, ws, interval_sec):
        headers = ["時間", "狀態", "讀取速度 (MB/Sec)",
                   "寫入速度 (MB/Sec)", "外殼溫度 (°C)", "SMART 溫度 (°C)", "採樣間隔 (秒)"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font      = Font(bold=True, color="FFFFFF", name="微軟正黑體")
            cell.fill      = PatternFill(fill_type="solid", fgColor=self.HDR_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = Border(bottom=Side(style="medium", color="FFFFFF"))
        ws.cell(row=2, column=7, value=interval_sec)
        # H 欄標頭：累積時間
        ws.cell(row=1, column=self.TIME_COL, value="累積時間 (s)")
        ws.row_dimensions[1].height = 24

    def _data(self, ws, records, interval_sec):
        thin = Side(style="thin", color="CCCCCC")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        STATUS_ZH = {"writing":   "寫入 Writing",
                     "verifying": "讀取 Verifying",
                     "unknown":   "偵測中"}
        for i, (ts, d) in enumerate(records, 2):
            ws.cell(row=i, column=1,
                    value=ts.strftime("%Y-%m-%d %H:%M:%S")).alignment = \
                Alignment(horizontal="center")
            ws.cell(row=i, column=2,
                    value=STATUS_ZH.get(d.get("status", "unknown"), ""))
            ws.cell(row=i, column=3,
                    value=d.get("read_speed")  if d.get("read_speed")  is not None else "")
            ws.cell(row=i, column=4,
                    value=d.get("write_speed") if d.get("write_speed") is not None else "")
            ws.cell(row=i, column=5,
                    value=d.get("temperature") if d.get("temperature") is not None else "")
            ws.cell(row=i, column=6,
                    value=d.get("smart_temp")  if d.get("smart_temp")  is not None else "")
            # H 欄：累積時間（秒）= (筆數-1) × 採樣間隔
            ws.cell(row=i, column=self.TIME_COL, value=(i - 2) * interval_sec)
            for c in range(1, 7):
                ws.cell(row=i, column=c).border = brd
                ws.cell(row=i, column=c).font   = Font(name="微軟正黑體")
            if i % 2 == 0:
                for c in range(1, 7):
                    ws.cell(row=i, column=c).fill = PatternFill(
                        fill_type="solid", fgColor="EEF2FF")

    def _make_scatter(self, title, y_title):
        """建立 ScatterChart，正確顯示 X/Y 軸刻度"""
        ch = ScatterChart()
        ch.title  = title
        ch.style  = 10
        ch.height = 18
        ch.width  = 34
        # ★ delete=False 是關鍵：不設此值 Excel 會隱藏軸刻度
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        ch.x_axis.title  = "時間 (秒)"
        ch.y_axis.title  = y_title
        ch.x_axis.numFmt = '0"s"'   # 顯示為 0s / 5s / 10s
        ch.y_axis.numFmt = "0.0"    # 顯示一位小數
        ch.legend.position = "t"
        return ch

    def _add_series(self, ch, ws, n, y_col, color, name):
        """加入一條折線 series"""
        xvals = Reference(ws, min_col=self.TIME_COL, min_row=2, max_row=n + 1)
        yvals = Reference(ws, min_col=y_col,         min_row=2, max_row=n + 1)
        ser = Series(yvals, xvals, title=name)
        ser.graphicalProperties.line.solidFill = color
        ser.graphicalProperties.line.width     = 22000
        ser.marker.symbol = "none"
        ch.series.append(ser)

    def _chart_speed(self, wb, ws, records):
        n = len(records)
        if n < 2:
            return
        cws = wb.create_sheet("速度圖表")
        ch  = self._make_scatter("磁碟讀寫速度趨勢", "Speed (MB/Sec)")
        self._add_series(ch, ws, n, 3, self.READ_COLOR,  "讀取速度 (MB/Sec)")
        self._add_series(ch, ws, n, 4, self.WRITE_COLOR, "寫入速度 (MB/Sec)")
        cws.add_chart(ch, "B2")

    def _chart_temp(self, wb, ws, records):
        has_shell = any(d.get("temperature") is not None for _, d in records)
        has_smart = any(d.get("smart_temp")  is not None for _, d in records)
        if (not has_shell and not has_smart) or len(records) < 2:
            return
        cws = wb.create_sheet("溫度圖表")
        ch  = self._make_scatter("溫度趨勢 (°C)", "Temperature (°C)")
        n   = len(records)
        if has_shell:
            self._add_series(ch, ws, n, 5, self.SHELL_TEMP_COLOR, "外殼溫度 (°C)")
        if has_smart:
            self._add_series(ch, ws, n, 6, self.SMART_TEMP_COLOR, "SMART 溫度 (°C)")
        cws.add_chart(ch, "B2")

    def _widths(self, ws):
        for col, w in zip("ABCDEFGH", [22, 18, 20, 20, 16, 16, 14, 14]):
            ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════
#  GUI 主介面
# ══════════════════════════════════════════════════════════════════
class App(tk.Tk):
    VERSION = "1.6.5"

    def __init__(self):
        super().__init__()
        self.title(f"BurnInTest 磁碟速度 + 溫度監控 v{self.VERSION}")
        self.resizable(True, True)
        self.minsize(980, 700)
        self.configure(bg="#F0F4FF")

        self.records        = []
        self.monitor_thread = None
        self._region        = None
        self._last_read     = None
        self._last_write    = None
        self.temp_reader    = TemperatureReader()   # 外殼溫度（序列埠）
        self.smart_manager  = SmartTempManager()    # SMART 溫度（USB SSD）

        self._check_deps()
        self._build_ui()
        self._apply_style()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── 依賴套件 ────────────────────────────────────────────────
    def _check_deps(self):
        if MISSING:
            messagebox.showerror("缺少套件",
                "請先安裝以下套件：\n\n  pip install " + " ".join(MISSING)
                + "\n\n以及 Tesseract OCR：\n"
                  "https://github.com/UB-Mannheim/tesseract/wiki")
            sys.exit(1)
        tess_exe = next(
            (p for p in TESSERACT_DEFAULT_PATHS if os.path.isfile(p)), None)
        if tess_exe:
            pytesseract.pytesseract.tesseract_cmd = tess_exe
        else:
            msg = (
                "找不到 Tesseract OCR 可執行檔。\n\n"
                "請選擇以下其中一種方式：\n"
                "  1. 點選『是』手動選擇 tesseract.exe 路徑\n"
                "  2. 點選『否』前往下載安裝後重新執行\n"
                "     https://github.com/UB-Mannheim/tesseract/wiki"
            )
            if messagebox.askyesno("找不到 Tesseract", msg):
                path = filedialog.askopenfilename(
                    title="選擇 tesseract.exe",
                    filetypes=[("Executable", "*.exe")])
                if path:
                    pytesseract.pytesseract.tesseract_cmd = path
                else:
                    messagebox.showwarning("警告", "未選擇 Tesseract 路徑，OCR 功能將無法使用。")
            else:
                import webbrowser
                webbrowser.open("https://github.com/UB-Mannheim/tesseract/wiki")
                messagebox.showinfo("提示", "請安裝 Tesseract 後重新執行程式。")
                sys.exit(0)

    # ── UI 建置 ─────────────────────────────────────────────────
    def _build_ui(self):
        self._build_disk_cfg()
        self._build_region_cfg()
        self._build_shell_temp_cfg()    # 外殼溫度（序列埠）
        self._build_smart_temp_cfg()    # SMART 溫度（USB SSD）
        self._build_buttons()
        self._build_realtime_display()
        self._build_table()

    # ── 磁碟監控設定 ──
    def _build_disk_cfg(self):
        cfg = tk.LabelFrame(self, text=" ⚙  磁碟監控設定 ", bg="#F0F4FF",
            fg="#2F5496", font=("微軟正黑體", 10, "bold"), padx=10, pady=8)
        cfg.pack(fill="x", padx=12, pady=(10, 4))

        tk.Label(cfg, text="採樣間隔（秒）：", bg="#F0F4FF",
            font=("微軟正黑體", 9)).grid(row=0, column=0, sticky="w", padx=4)
        self.var_interval = tk.StringVar(value="5")
        ttk.Spinbox(cfg, from_=1, to=3600, textvariable=self.var_interval,
            width=8, font=("微軟正黑體", 9)).grid(row=0, column=1, padx=4)

        tk.Label(cfg, text="監控時長（分鐘，0=無限）：", bg="#F0F4FF",
            font=("微軟正黑體", 9)).grid(row=0, column=2, sticky="w", padx=12)
        self.var_duration = tk.StringVar(value="60")
        ttk.Spinbox(cfg, from_=0, to=9999, textvariable=self.var_duration,
            width=8, font=("微軟正黑體", 9)).grid(row=0, column=3, padx=4)

        tk.Label(cfg, text="擷取螢幕：", bg="#F0F4FF",
            font=("微軟正黑體", 9)).grid(row=0, column=4, sticky="w", padx=12)
        self.var_monitor = tk.StringVar(value="全螢幕")
        ttk.Combobox(cfg, textvariable=self.var_monitor, width=10,
            font=("微軟正黑體", 9), state="readonly",
            values=self._get_monitor_names()).grid(row=0, column=5, padx=4)

        mode_txt   = "擷取模式：視窗直讀 ✓（已安裝 pywinauto）" if HAS_PYWINAUTO \
                     else "擷取模式：OCR 截圖（建議安裝 pywinauto 提升準確度）"
        mode_color = "#375623" if HAS_PYWINAUTO else "#C55A11"
        self.lbl_mode = tk.Label(cfg, text=mode_txt, bg="#F0F4FF",
            fg=mode_color, font=("微軟正黑體", 9))
        self.lbl_mode.grid(row=1, column=0, columnspan=6,
            sticky="w", padx=4, pady=(4, 0))

    # ── OCR 截取區域 ──
    def _build_region_cfg(self):
        reg = tk.LabelFrame(self, text=" 🖱  截取區域設定（OCR 模式使用）",
            bg="#F0F4FF", fg="#2F5496",
            font=("微軟正黑體", 10, "bold"), padx=10, pady=6)
        reg.pack(fill="x", padx=12, pady=(0, 4))

        tk.Button(reg, text="🖱  滑鼠拖拉選取區域",
            command=self._open_region_selector,
            bg="#5B4DA0", fg="white", font=("微軟正黑體", 9, "bold"),
            relief="flat", padx=14, pady=4, cursor="hand2",
            activebackground="#3E3472", activeforeground="white"
            ).grid(row=0, column=0, padx=4, pady=2)

        tk.Button(reg, text="✕  使用全螢幕",
            command=self._clear_region,
            bg="#7F7F7F", fg="white", font=("微軟正黑體", 9),
            relief="flat", padx=12, pady=4, cursor="hand2",
            activebackground="#595959", activeforeground="white"
            ).grid(row=0, column=1, padx=4, pady=2)

        self.lbl_region = tk.Label(reg,
            text="目前：全螢幕（視窗直讀模式不需設定）",
            bg="#F0F4FF", fg="#666", font=("微軟正黑體", 9))
        self.lbl_region.grid(row=0, column=2, padx=8, sticky="w")

        for i, (lbl, attr) in enumerate([("X:", "entry_rx"), ("Y:", "entry_ry"),
                                          ("W:", "entry_rw"), ("H:", "entry_rh")]):
            tk.Label(reg, text=lbl, bg="#F0F4FF",
                font=("微軟正黑體", 9)).grid(row=0, column=3+i*2,
                padx=(8 if i == 0 else 2, 2))
            e = ttk.Entry(reg, width=6, font=("微軟正黑體", 9))
            e.grid(row=0, column=4+i*2, padx=2)
            setattr(self, attr, e)

        tk.Button(reg, text="套用", command=self._apply_manual_region,
            bg="#2F5496", fg="white", font=("微軟正黑體", 9),
            relief="flat", padx=8, pady=2, cursor="hand2"
            ).grid(row=0, column=12, padx=6)

    # ── 外殼溫度感測器設定（序列埠）──
    def _build_shell_temp_cfg(self):
        lf = tk.LabelFrame(self, text=" 🌡  外殼溫度感測器設定（序列埠）",
            bg="#F0F4FF", fg="#C00000",
            font=("微軟正黑體", 10, "bold"), padx=10, pady=8)
        lf.pack(fill="x", padx=12, pady=(0, 4))

        self.var_temp_enable = tk.BooleanVar(value=False)
        ttk.Checkbutton(lf, text="啟用外殼溫度",
            variable=self.var_temp_enable,
            command=self._on_temp_toggle
            ).grid(row=0, column=0, padx=4, sticky="w")

        tk.Label(lf, text="序列埠：", bg="#F0F4FF",
            font=("微軟正黑體", 9)).grid(row=0, column=1, padx=(12,2), sticky="w")
        self.var_port = tk.StringVar(value="")
        self.port_combo = ttk.Combobox(lf, textvariable=self.var_port,
            width=12, font=("微軟正黑體", 9), state="readonly")
        self.port_combo.grid(row=0, column=2, padx=4)

        tk.Button(lf, text="🔄 刷新",
            command=self._refresh_ports,
            bg="#4A6FA5", fg="white", font=("微軟正黑體", 9),
            relief="flat", padx=10, pady=3, cursor="hand2",
            activebackground="#3A5A8A", activeforeground="white"
            ).grid(row=0, column=3, padx=4)

        self.btn_connect = tk.Button(lf, text="🔌 連接",
            command=self._connect_serial,
            bg="#375623", fg="white", font=("微軟正黑體", 9, "bold"),
            relief="flat", padx=12, pady=3, cursor="hand2",
            activebackground="#274018", activeforeground="white")
        self.btn_connect.grid(row=0, column=4, padx=4)

        self.btn_disconnect = tk.Button(lf, text="⛔ 斷線",
            command=self._disconnect_serial,
            bg="#C55A11", fg="white", font=("微軟正黑體", 9),
            relief="flat", padx=12, pady=3, cursor="hand2", state="disabled",
            activebackground="#9E4710", activeforeground="white")
        self.btn_disconnect.grid(row=0, column=5, padx=4)

        proto_txt = f"協定：baudrate={TEMP_BAUDRATE}  指令=AA 55 01 03 03  回應第4~5 byte ÷10 = 外殼溫度(°C)"
        tk.Label(lf, text=proto_txt, bg="#F0F4FF",
            fg="#888", font=("微軟正黑體", 8)).grid(
            row=1, column=0, columnspan=8, sticky="w", padx=4, pady=(2,0))

        self.lbl_temp_status = tk.Label(lf,
            text="外殼溫度感測器：未連接" if HAS_SERIAL else "外殼溫度感測器：需安裝 pyserial",
            bg="#F0F4FF",
            fg="#666" if HAS_SERIAL else "#C00000",
            font=("微軟正黑體", 9))
        self.lbl_temp_status.grid(row=0, column=6, padx=16, sticky="w")

        self._refresh_ports()

    # ── SMART 溫度設定（USB SSD）──
    def _build_smart_temp_cfg(self):
        lf = tk.LabelFrame(self, text=" 💾  USB SSD SMART 溫度設定",
            bg="#F0F4FF", fg="#7030A0",
            font=("微軟正黑體", 10, "bold"), padx=10, pady=8)
        lf.pack(fill="x", padx=12, pady=(0, 4))

        # 管理員狀態提示
        is_admin = _is_admin()
        admin_txt   = "✓ 已取得管理員權限" if is_admin else "⚠ 未取得管理員權限（SMART 溫度讀取需要）"
        admin_color = "#375623" if is_admin else "#C55A11"
        tk.Label(lf, text=admin_txt, bg="#F0F4FF", fg=admin_color,
            font=("微軟正黑體", 8)).grid(row=0, column=0, sticky="w", padx=4)

        # 啟用勾選框
        self.var_smart_enable = tk.BooleanVar(value=False)
        ttk.Checkbutton(lf, text="啟用 SMART 溫度",
            variable=self.var_smart_enable,
            command=self._on_smart_toggle
            ).grid(row=0, column=1, padx=(12, 4), sticky="w")

        # 裝置選擇
        tk.Label(lf, text="USB SSD：", bg="#F0F4FF",
            font=("微軟正黑體", 9)).grid(row=0, column=2, padx=(12,2), sticky="w")
        self.var_smart_drive = tk.StringVar(value="")
        self.smart_combo = ttk.Combobox(lf, textvariable=self.var_smart_drive,
            width=32, font=("微軟正黑體", 9), state="readonly")
        self.smart_combo.grid(row=0, column=3, padx=4)
        self.smart_combo.bind("<<ComboboxSelected>>", self._on_smart_drive_select)

        # 掃描按鈕
        self.btn_smart_scan = tk.Button(lf, text="🔍 掃描裝置",
            command=self._scan_smart_drives,
            bg="#7030A0", fg="white", font=("微軟正黑體", 9, "bold"),
            relief="flat", padx=12, pady=3, cursor="hand2",
            activebackground="#4B1070", activeforeground="white")
        self.btn_smart_scan.grid(row=0, column=4, padx=4)

        # 狀態標籤
        self.lbl_smart_status = tk.Label(lf,
            text="SMART 溫度：未啟用",
            bg="#F0F4FF", fg="#666", font=("微軟正黑體", 9))
        self.lbl_smart_status.grid(row=0, column=5, padx=16, sticky="w")

        # 讀取方式說明
        self.lbl_smart_method = tk.Label(lf,
            text="讀取方式：-",
            bg="#F0F4FF", fg="#888", font=("微軟正黑體", 8))
        self.lbl_smart_method.grid(
            row=1, column=0, columnspan=8, sticky="w", padx=4, pady=(2,0))

        # 內部裝置清單
        self._smart_drives: list = []

    # ── 控制按鈕 ──
    def _build_buttons(self):
        btn = tk.Frame(self, bg="#F0F4FF")
        btn.pack(fill="x", padx=12, pady=4)

        self.btn_start = tk.Button(btn, text="▶  開始監控",
            command=self.start_monitor, bg="#2F5496", fg="white",
            font=("微軟正黑體", 10, "bold"), relief="flat",
            padx=20, pady=6, cursor="hand2",
            activebackground="#1F3D80", activeforeground="white")
        self.btn_start.pack(side="left", padx=4)

        self.btn_stop = tk.Button(btn, text="■  停止監控",
            command=self.stop_monitor, bg="#C55A11", fg="white",
            font=("微軟正黑體", 10, "bold"), relief="flat",
            padx=20, pady=6, cursor="hand2", state="disabled",
            activebackground="#9E4710", activeforeground="white")
        self.btn_stop.pack(side="left", padx=4)

        self.btn_export = tk.Button(btn, text="💾  匯出 Excel",
            command=self.export_excel, bg="#375623", fg="white",
            font=("微軟正黑體", 10, "bold"), relief="flat",
            padx=20, pady=6, cursor="hand2", state="disabled",
            activebackground="#274018", activeforeground="white")
        self.btn_export.pack(side="left", padx=4)

        self.btn_clear = tk.Button(btn, text="🗑  清除資料",
            command=self.clear_data, bg="#7F7F7F", fg="white",
            font=("微軟正黑體", 10), relief="flat",
            padx=16, pady=6, cursor="hand2",
            activebackground="#595959", activeforeground="white")
        self.btn_clear.pack(side="left", padx=4)

        self.lbl_status = tk.Label(btn, text="● 就緒", bg="#F0F4FF",
            fg="#666", font=("微軟正黑體", 9))
        self.lbl_status.pack(side="right", padx=12)

    # ── 即時數值顯示列 ──
    def _build_realtime_display(self):
        spd = tk.Frame(self, bg="#1A1A2E")
        spd.pack(fill="x", padx=12, pady=4)

        # 狀態
        tk.Label(spd, text="狀態：", bg="#1A1A2E", fg="#AAAAAA",
            font=("微軟正黑體", 9)).grid(row=0, column=0, padx=(14,2), pady=10)
        self.lbl_state = tk.Label(spd, text="--", bg="#1A1A2E", fg="#FFFFFF",
            font=("微軟正黑體", 11, "bold"), width=16, anchor="w")
        self.lbl_state.grid(row=0, column=1, padx=(2,12))

        # 讀取速度
        tk.Label(spd, text="讀取：", bg="#1A1A2E", fg="#88B4FF",
            font=("微軟正黑體", 9)).grid(row=0, column=2, padx=(0,2))
        self.lbl_read = tk.Label(spd, text="-- MB/s", bg="#1A1A2E", fg="#4FC3F7",
            font=("Consolas", 17, "bold"), width=12)
        self.lbl_read.grid(row=0, column=3, padx=(2,12))

        # 寫入速度
        tk.Label(spd, text="寫入：", bg="#1A1A2E", fg="#FFAB76",
            font=("微軟正黑體", 9)).grid(row=0, column=4, padx=(0,2))
        self.lbl_write = tk.Label(spd, text="-- MB/s", bg="#1A1A2E", fg="#FFB74D",
            font=("Consolas", 17, "bold"), width=12)
        self.lbl_write.grid(row=0, column=5, padx=(2,12))

        # 外殼溫度
        tk.Label(spd, text="外殼：", bg="#1A1A2E", fg="#FF8A80",
            font=("微軟正黑體", 9)).grid(row=0, column=6, padx=(0,2))
        self.lbl_temp = tk.Label(spd, text="-- °C", bg="#1A1A2E", fg="#FF5252",
            font=("Consolas", 17, "bold"), width=9)
        self.lbl_temp.grid(row=0, column=7, padx=(2,12))

        # SMART 溫度
        tk.Label(spd, text="SMART：", bg="#1A1A2E", fg="#CE93D8",
            font=("微軟正黑體", 9)).grid(row=0, column=8, padx=(0,2))
        self.lbl_smart_temp = tk.Label(spd, text="-- °C", bg="#1A1A2E", fg="#BA68C8",
            font=("Consolas", 17, "bold"), width=9)
        self.lbl_smart_temp.grid(row=0, column=9, padx=(2,12))

        # 已記錄筆數
        tk.Label(spd, text="已記錄：", bg="#1A1A2E", fg="#AAAAAA",
            font=("微軟正黑體", 9)).grid(row=0, column=10, padx=(0,2))
        self.lbl_count = tk.Label(spd, text="0 筆", bg="#1A1A2E", fg="#A5D6A7",
            font=("Consolas", 15, "bold"))
        self.lbl_count.grid(row=0, column=11, padx=(2,14))

    # ── 資料表格 ──
    def _build_table(self):
        tbl = tk.Frame(self, bg="#F0F4FF")
        tbl.pack(fill="both", expand=True, padx=12, pady=(4, 8))

        cols = ("時間", "狀態", "讀取速度 (MB/Sec)",
                "寫入速度 (MB/Sec)", "外殼溫度 (°C)", "SMART 溫度 (°C)")
        self.tree = ttk.Treeview(tbl, columns=cols, show="headings", height=10)
        for col, w in zip(cols, [180, 150, 150, 150, 120, 120]):
            self.tree.heading(col, text=col)
            self.tree.column(col, anchor="center", width=w)
        self.tree.tag_configure("writing",   background="#FFF3E0")
        self.tree.tag_configure("verifying", background="#E3F2FD")
        self.tree.tag_configure("unknown",   background="#F5F5F5")

        sy = ttk.Scrollbar(tbl, orient="vertical",   command=self.tree.yview)
        sx = ttk.Scrollbar(tbl, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")
        tbl.rowconfigure(0, weight=1)
        tbl.columnconfigure(0, weight=1)

    def _apply_style(self):
        s = ttk.Style()
        s.theme_use("clam")
        s.configure("Treeview",
            font=("微軟正黑體", 9), rowheight=24,
            background="white", fieldbackground="white")
        s.configure("Treeview.Heading",
            font=("微軟正黑體", 9, "bold"),
            background="#2F5496", foreground="white")
        s.map("Treeview", background=[("selected", "#BDD7EE")])

    # ── 外殼溫度序列埠操作 ───────────────────────────────────────
    def _refresh_ports(self):
        ports = TemperatureReader.list_ports()
        self.port_combo["values"] = ports
        if ports:
            if self.var_port.get() not in ports:
                self.var_port.set(ports[0])
        else:
            self.var_port.set("")
        if not HAS_SERIAL:
            self.lbl_temp_status.configure(
                text="外殼溫度感測器：需安裝 pyserial（pip install pyserial）",
                fg="#C00000")

    def _on_temp_toggle(self):
        if not self.var_temp_enable.get():
            self._disconnect_serial()

    def _connect_serial(self):
        if not HAS_SERIAL:
            messagebox.showerror("錯誤",
                "未安裝 pyserial，請執行：\npip install pyserial")
            return
        port = self.var_port.get().strip()
        if not port:
            messagebox.showerror("錯誤", "請先選擇序列埠。")
            return

        self.btn_connect.configure(state="disabled", text="🔌 連接中…")
        self.lbl_temp_status.configure(text=f"正在連接 {port}…", fg="#888")
        self.update_idletasks()

        def _do_connect():
            try:
                self.temp_reader.open(port)
                self.after(0, self._on_connect_success, port)
            except Exception as e:
                self.after(0, self._on_connect_fail, str(e))

        threading.Thread(target=_do_connect, daemon=True).start()

    def _on_connect_success(self, port):
        self.var_temp_enable.set(True)
        self.btn_connect.configure(state="disabled", text="🔌 連接")
        self.btn_disconnect.configure(state="normal")
        self.lbl_temp_status.configure(
            text=f"外殼溫度感測器：已連接 {port}（baudrate={TEMP_BAUDRATE}）",
            fg="#375623")

    def _on_connect_fail(self, err_msg):
        self.btn_connect.configure(state="normal", text="🔌 連接")
        self.lbl_temp_status.configure(text="外殼溫度感測器：連接失敗", fg="#C00000")
        messagebox.showerror("連接失敗", err_msg)

    def _disconnect_serial(self):
        self.temp_reader.close()
        self.var_temp_enable.set(False)
        self.btn_connect.configure(state="normal")
        self.btn_disconnect.configure(state="disabled")
        self.lbl_temp_status.configure(text="外殼溫度感測器：已斷線", fg="#666")
        self.lbl_temp.configure(text="-- °C")

    # ── SMART 溫度操作 ────────────────────────────────────────────
    def _on_smart_toggle(self):
        enabled = self.var_smart_enable.get()
        self.smart_manager.enabled = enabled
        if not enabled:
            self.lbl_smart_status.configure(text="SMART 溫度：已停用", fg="#666")
            self.lbl_smart_temp.configure(text="-- °C")

    def _scan_smart_drives(self):
        self.btn_smart_scan.configure(state="disabled", text="🔍 掃描中…")
        self.lbl_smart_status.configure(text="掃描 USB 裝置中…", fg="#888")
        self.update_idletasks()

        def _do_scan():
            drives = self.smart_manager.scan_drives()
            self.after(0, self._on_scan_done, drives)

        threading.Thread(target=_do_scan, daemon=True).start()

    def _on_scan_done(self, drives: list):
        self.btn_smart_scan.configure(state="normal", text="🔍 掃描裝置")
        self._smart_drives = drives
        if not drives:
            self.smart_combo["values"] = []
            self.var_smart_drive.set("")
            self.lbl_smart_status.configure(
                text="SMART 溫度：未偵測到 USB SSD（需管理員權限）", fg="#C55A11")
            return
        labels = [f"{d.model}  [{d.device_path}]" for d in drives]
        self.smart_combo["values"] = labels
        self.var_smart_drive.set(labels[0])
        self.smart_manager.select_drive(drives[0])
        self.lbl_smart_status.configure(
            text=f"SMART 溫度：偵測到 {len(drives)} 台 USB SSD", fg="#375623")

    def _on_smart_drive_select(self, event=None):
        idx = self.smart_combo.current()
        if 0 <= idx < len(self._smart_drives):
            self.smart_manager.select_drive(self._smart_drives[idx])
            d = self._smart_drives[idx]
            self.lbl_smart_status.configure(
                text=f"SMART 溫度：已選擇 {d.model}", fg="#375623")

    # ── 截取區域 ────────────────────────────────────────────────
    def _open_region_selector(self):
        self.iconify()
        self.after(300, self._do_select)

    def _do_select(self):
        r = RegionSelector(self).select()
        self.deiconify()
        if r:
            self._region = r
            self.lbl_region.configure(
                text=f"已選取：X={r['x']}  Y={r['y']}  W={r['w']}  H={r['h']}",
                fg="#375623")
            for e, k in zip(
                    (self.entry_rx, self.entry_ry, self.entry_rw, self.entry_rh),
                    ("x","y","w","h")):
                e.delete(0, tk.END); e.insert(0, str(r[k]))
        else:
            messagebox.showinfo("提示", "未選取區域，將使用全螢幕。")

    def _clear_region(self):
        self._region = None
        self.lbl_region.configure(
            text="目前：全螢幕（視窗直讀模式不需設定）", fg="#666")
        for e in (self.entry_rx, self.entry_ry, self.entry_rw, self.entry_rh):
            e.delete(0, tk.END)

    def _apply_manual_region(self):
        try:
            r = {k: int(e.get()) for k, e in zip(
                ("x","y","w","h"),
                (self.entry_rx, self.entry_ry, self.entry_rw, self.entry_rh))}
            if r["w"] <= 0 or r["h"] <= 0:
                raise ValueError
            self._region = r
            self.lbl_region.configure(
                text=f"已設定：X={r['x']}  Y={r['y']}  W={r['w']}  H={r['h']}",
                fg="#375623")
        except (ValueError, tk.TclError):
            messagebox.showerror("錯誤", "請輸入有效整數（W/H 需大於 0）。")

    # ── 輔助 ────────────────────────────────────────────────────
    def _get_monitor_names(self):
        try:
            with mss_module.mss() as sct:
                return ["全螢幕"] + [f"螢幕 {i}" for i in range(1, len(sct.monitors))]
        except Exception:
            return ["全螢幕"]

    def _update_status(self, text, color="#666"):
        self.lbl_status.configure(text=text, fg=color)

    # ── 監控啟停 ────────────────────────────────────────────────
    def start_monitor(self):
        try:
            interval     = int(self.var_interval.get())
            duration_min = int(self.var_duration.get())
        except ValueError:
            messagebox.showerror("錯誤", "請輸入有效的數字。"); return
        if interval < 1:
            messagebox.showerror("錯誤", "採樣間隔至少需 1 秒。"); return

        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._update_status("● 監控中…", "#2F5496")

        self.monitor_thread = MonitorThread(
            interval_sec=interval,
            total_sec=duration_min * 60,
            region=self._region,
            temp_reader=self.temp_reader,
            smart_manager=self.smart_manager,
            on_record=self._on_record,
            on_finish=self._on_finish,
            on_error=self._on_error,
            on_mode_update=self._on_mode_update,
        )
        self.monitor_thread.start()

    def stop_monitor(self):
        if self.monitor_thread:
            self.monitor_thread.stop()
        self.btn_stop.configure(state="disabled")
        self._update_status("● 停止中…", "#C55A11")

    # ── 資料回呼 ────────────────────────────────────────────────
    def _on_record(self, ts, data):
        read_spd  = data.get("read_speed")
        write_spd = data.get("write_speed")

        if read_spd is not None:
            self._last_read = read_spd
            if write_spd is None:
                write_spd = self._last_write
        elif write_spd is not None:
            self._last_write = write_spd
            if read_spd is None:
                read_spd = self._last_read

        filled = dict(data)
        filled["read_speed"]  = read_spd
        filled["write_speed"] = write_spd

        self.records.append((ts, filled))
        self.after(0, self._update_ui, ts, filled)

    def _update_ui(self, ts, data):
        status     = data.get("status",      "unknown")
        read_spd   = data.get("read_speed")
        write_spd  = data.get("write_speed")
        temp       = data.get("temperature")   # 外殼溫度
        smart_temp = data.get("smart_temp")    # SMART 溫度

        STATE_LABELS = {
            "writing":   ("寫入中 (Writing)",   "#FFB74D"),
            "verifying": ("讀取中 (Verifying)", "#4FC3F7"),
            "unknown":   ("偵測中…",            "#AAAAAA"),
        }
        lbl, col = STATE_LABELS.get(status, ("未知", "#FFF"))
        self.lbl_state.configure(text=lbl, fg=col)

        if read_spd   is not None:
            self.lbl_read.configure(text=f"{read_spd:.1f} MB/s")
        if write_spd  is not None:
            self.lbl_write.configure(text=f"{write_spd:.1f} MB/s")
        if temp       is not None:
            self.lbl_temp.configure(text=f"{temp:.1f} °C")
        if smart_temp is not None:
            self.lbl_smart_temp.configure(text=f"{smart_temp} °C")
            # 同步更新 SMART 讀取方式說明
            d = self.smart_manager.selected_drive
            if d and d.method:
                self.lbl_smart_method.configure(text=f"讀取方式：{d.method}")

        self.lbl_count.configure(text=f"{len(self.records)} 筆")

        STATUS_ZH = {"writing":   "寫入 Writing",
                     "verifying": "讀取 Verifying",
                     "unknown":   "偵測中"}
        self.tree.insert("", "end", tags=(status,), values=(
            ts.strftime("%Y-%m-%d %H:%M:%S"),
            STATUS_ZH.get(status, status),
            f"{read_spd:.1f}"   if read_spd   is not None else "",
            f"{write_spd:.1f}"  if write_spd  is not None else "",
            f"{temp:.1f}"       if temp        is not None else "",
            f"{smart_temp}"     if smart_temp  is not None else "",
        ))
        self.tree.yview_moveto(1.0)
        if self.records:
            self.btn_export.configure(state="normal")

    def _on_mode_update(self, mode):
        self.after(0, lambda: self.lbl_mode.configure(
            text=f"擷取模式：{mode}",
            fg="#375623" if "✓" in mode else "#C55A11"))

    def _on_finish(self):
        self.after(0, self._finish_ui)

    def _finish_ui(self):
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self._update_status(f"● 完成（共 {len(self.records)} 筆）", "#375623")
        messagebox.showinfo("完成",
            f"監控結束，共記錄 {len(self.records)} 筆。\n"
            "可按「匯出 Excel」儲存報表。")

    def _on_error(self, msg):
        self.after(0, lambda: self._update_status(
            f"⚠ 錯誤：{msg[:40]}", "#C00000"))

    # ── 資料操作 ────────────────────────────────────────────────
    def clear_data(self):
        if self.records and not messagebox.askyesno(
                "確認", "確定清除所有已記錄的資料？"):
            return
        self.records.clear()
        self._last_read  = None
        self._last_write = None
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.lbl_state.configure(text="--",      fg="#FFFFFF")
        self.lbl_read.configure( text="-- MB/s")
        self.lbl_write.configure(text="-- MB/s")
        self.lbl_temp.configure( text="-- °C")
        self.lbl_smart_temp.configure(text="-- °C")
        self.lbl_count.configure(text="0 筆")
        self.btn_export.configure(state="disabled")
        self._update_status("● 就緒", "#666")

    def export_excel(self):
        if not self.records:
            messagebox.showwarning("警告", "沒有資料可匯出。"); return
        name = datetime.datetime.now().strftime("BurnIn_速度報表_%Y%m%d_%H%M%S.xlsx")
        path = filedialog.asksaveasfilename(
            title="儲存 Excel 報表", defaultextension=".xlsx",
            initialfile=name,
            filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")])
        if not path: return
        try:
            interval = int(self.var_interval.get())
        except ValueError:
            interval = 0
        try:
            ExcelExporter().export(self.records, path, interval)
            messagebox.showinfo("匯出成功", f"已儲存至：\n{path}")
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("匯出失敗", str(e))

    def _on_close(self):
        """視窗關閉時清理資源"""
        try:
            if self.monitor_thread:
                self.monitor_thread.stop()
        except Exception:
            pass
        self.temp_reader.close()
        self.destroy()


# ══════════════════════════════════════════════════════════════════
#  程式入口（自動 UAC 提升）
# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    if sys.platform == "win32" and not _is_admin():
        _relaunch_as_admin()
        sys.exit(0)
    App().mainloop()
